"""PII detection engine using regex patterns for Indian and global sensitive data.

Provides regex-based detection for Aadhaar, PAN, Email, Phone (India),
and Credit Card numbers with Luhn algorithm validation. Supports CSV, XLSX,
PDF, and TXT file types.
"""

import re
import csv
import io
import os
from typing import Any, Optional

from app.utils.exceptions import ValidationException

# ---------------------------------------------------------------------------
# Regex patterns
# ---------------------------------------------------------------------------

PATTERNS = {
    "Aadhaar": {
        "regex": re.compile(r"\b[2-9]\d{11}\b"),
        "severity": "CRITICAL",
    },
    "PAN": {
        "regex": re.compile(r"[A-Z]{5}[0-9]{4}[A-Z]{1}"),
        "severity": "HIGH",
    },
    "Email": {
        "regex": re.compile(r"[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}"),
        "severity": "LOW",
    },
    "Phone": {
        "regex": re.compile(r"(?:\+91|0)?[6-9]\d{9}"),
        "severity": "MEDIUM",
    },
    "CreditCard": {
        "regex": re.compile(r"\b(?:\d[ -]*?){13,16}\b"),
        "severity": "CRITICAL",
    },
    "Passport": {
        "regex": re.compile(r"\b[A-Z][0-9]{7}\b"),
        "severity": "HIGH",
    },
    "DOB": {
        "regex": re.compile(r"\b\d{2}[/-]\d{2}[/-]\d{4}\b"),
        "severity": "MEDIUM",
    },
}


def _luhn_check(card_number: str) -> bool:
    """Validate a credit card number using the Luhn algorithm.

    Strips non-digit characters, then applies the standard Luhn checksum.
    Returns True if the number passes validation.
    """
    digits = [int(ch) for ch in card_number if ch.isdigit()]
    if len(digits) < 13 or len(digits) > 16:
        return False

    # Luhn: double every second digit from the right
    checksum = 0
    double = False
    for d in reversed(digits):
        if double:
            d *= 2
            if d > 9:
                d -= 9
        checksum += d
        double = not double

    return checksum % 10 == 0


def _filter_credit_cards(candidates: list[str]) -> list[str]:
    """Filter raw regex matches to those that pass Luhn validation."""
    valid: list[str] = []
    for raw in candidates:
        cleaned = re.sub(r"[ -]", "", raw)
        if _luhn_check(cleaned):
            valid.append(cleaned)
    return valid


def _deduplicate_preserve_order(items: list[str]) -> list[str]:
    """Remove duplicates while preserving insertion order."""
    seen: set[str] = set()
    result: list[str] = []
    for item in items:
        if item not in seen:
            seen.add(item)
            result.append(item)
    return result


# ---------------------------------------------------------------------------
# Severity mapping for external consumers
# ---------------------------------------------------------------------------

SEVERITY_MAP: dict[str, str] = {
    name: info["severity"] for name, info in PATTERNS.items()
}


class DetectionEngine:
    """Engine for detecting sensitive PII in text and files.

    Supports multiple file formats (CSV, XLSX, PDF, TXT) and returns
    structured findings with counts, sample values, and severity levels.
    """

    def detect(self, text: str) -> list[dict[str, Any]]:
        """Run all regex patterns against the provided text.

        Args:
            text: The raw text content to scan.

        Returns:
            A list of dicts, each containing:
                - data_type (str): Name of the detected PII type
                - count (int): Number of unique matches found
                - sample_values (list[str] | None): Up to 5 sample matches
                - severity (str): Severity level string
        """
        findings: list[dict[str, Any]] = []

        for data_type, pattern_info in PATTERNS.items():
            matches: list[str] = pattern_info["regex"].findall(text)

            if not matches:
                continue

            # Clean and validate matches
            if data_type == "CreditCard":
                valid_matches = _filter_credit_cards(matches)
                if not valid_matches:
                    continue
                unique_matches = _deduplicate_preserve_order(valid_matches)
            else:
                unique_matches = _deduplicate_preserve_order(matches)

            findings.append(
                {
                    "data_type": data_type,
                    "count": len(unique_matches),
                    "sample_values": unique_matches[:5],
                    "severity": pattern_info["severity"],
                }
            )

        return findings

    def detect_file(self, file_path: str, file_type: str) -> list[dict[str, Any]]:
        """Detect PII in a file by reading its contents based on file type.

        Args:
            file_path: Absolute or relative path to the file on disk.
            file_type: One of 'csv', 'xlsx', 'pdf', 'txt'.

        Returns:
            A list of detection finding dicts as returned by self.detect().

        Raises:
            ValidationException: If the file type is unsupported.
            FileNotFoundError: If the file does not exist.
        """
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"File not found: {file_path}")

        file_type = file_type.lower().lstrip(".")

        if file_type == "csv":
            text = self._detect_csv(file_path)
        elif file_type == "xlsx":
            text = self._detect_xlsx(file_path)
        elif file_type == "pdf":
            text = self._detect_pdf(file_path)
        elif file_type == "txt":
            text = self._detect_txt(file_path)
        else:
            raise ValidationException(
                {"file_type": [f"Unsupported file type: {file_type}"]}
            )

        return self.detect(text)

    def _detect_csv(self, file_path: str) -> str:
        """Parse a CSV file and concatenate all cell text for scanning."""
        text_parts: list[str] = []
        try:
            with open(file_path, encoding="utf-8", errors="replace") as f:
                reader = csv.reader(f)
                for row in reader:
                    text_parts.extend(cell.strip() for cell in row if cell.strip())
        except (csv.Error, IOError) as e:
            raise ValidationException({"file": [f"Error reading CSV: {str(e)}"]})
        return " ".join(text_parts)

    def _detect_xlsx(self, file_path: str) -> str:
        """Parse an XLSX file using openpyxl and concatenate cell text."""
        try:
            import openpyxl
        except ImportError:
            raise ImportError(
                "openpyxl is required for XLSX detection. "
                "Install it with: pip install openpyxl"
            )

        text_parts: list[str] = []
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                for row in ws.iter_rows(values_only=True):
                    for cell in row:
                        if cell is not None:
                            text_parts.append(str(cell).strip())
            wb.close()
        except Exception as e:
            raise ValidationException({"file": [f"Error reading XLSX: {str(e)}"]})
        return " ".join(text_parts)

    def _detect_pdf(self, file_path: str) -> str:
        """Parse a PDF file using PyPDF2 and concatenate all page text."""
        try:
            from PyPDF2 import PdfReader
        except ImportError:
            raise ImportError(
                "PyPDF2 is required for PDF detection. "
                "Install it with: pip install PyPDF2"
            )

        text_parts: list[str] = []
        try:
            reader = PdfReader(file_path)
            for page in reader.pages:
                page_text = page.extract_text()
                if page_text:
                    text_parts.append(page_text)
        except Exception as e:
            raise ValidationException({"file": [f"Error reading PDF: {str(e)}"]})
        return "\n".join(text_parts)

    def _detect_txt(self, file_path: str) -> str:
        """Read a plain text file and return its contents."""
        try:
            with open(file_path, encoding="utf-8", errors="replace") as f:
                return f.read()
        except IOError as e:
            raise ValidationException({"file": [f"Error reading TXT: {str(e)}"]})
